import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import tempfile
import os

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule


# === Helper: Determine scale and label based on max value in dataframe ===
def determine_scale_and_label(df):
    vals = []
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            vals.append(df[col].abs())
    if vals:
        vals_concat = pd.concat(vals)
        max_val = vals_concat.max()
        if max_val >= 1e7:
            return 1e7, "Crores INR (₹)"
        elif max_val >= 1e5:
            return 1e5, "Lakhs INR (₹)"
        else:
            return 1, "INR (₹)"
    else:
        return 1, "₹"

# === Fetch Historical Market Data ===
def fetch_data(ticker, period='1y', interval='1d'):
    data = yf.download(ticker, period=period, interval=interval)
    if isinstance(data.columns, pd.MultiIndex):
        data.columns = data.columns.droplevel(1)
    if data.empty:
        raise ValueError("No data fetched, check the ticker symbol or your internet connection!")
    return data

# === Calculate Technical Indicators ===
def calculate_indicators(df):
    if 'Close' not in df.columns:
        raise ValueError("Missing 'Close' column in data.")
    df['MA50'] = df['Close'].rolling(window=50).mean()
    df['MA200'] = df['Close'].rolling(window=200).mean()

    delta = df['Close'].diff()
    up = delta.clip(lower=0)
    down = -1 * delta.clip(upper=0)

    rsi_period = 14
    gain = up.rolling(window=rsi_period).mean()
    loss = down.rolling(window=rsi_period).mean()
    rs = gain / loss
    df['RSI'] = 100 - (100 / (1 + rs))
    return df

# === Investment Decision Logic ===
def investment_decision(df):
    latest = df.iloc[-1]
    if latest['MA50'] > latest['MA200'] and latest['RSI'] < 70:
        return "BUY"
    elif latest['RSI'] > 70:
        return "SELL"
    else:
        return "HOLD"

# === Fetch Financial Statements ===
def fetch_financial_statements(ticker):
    ticker_obj = yf.Ticker(ticker)
    try:
        bs = ticker_obj.balance_sheet
        inc = ticker_obj.financials
        cf = ticker_obj.cashflow
        return bs, inc, cf
    except Exception as e:
        print(f"Warning: Could not fetch financial statements for {ticker}: {e}")
        return None, None, None

# === Fetch Today's Market Cap ===
def get_market_cap(ticker):
    try:
        ticker_obj = yf.Ticker(ticker)
        mc = ticker_obj.info.get("marketCap", None)
        return mc
    except Exception as e:
        print(f"Could not fetch market cap: {e}")
        return None

# === Perform DCF Analysis ===
def perform_dcf_analysis(cf_df, years=5, growth_rate=0.05, discount_rate=0.10, terminal_growth=0.025):
    try:
        op_cf_keys = ['Total Cash From Operating Activities',
                      'Operating Cash Flow',
                      'Cash from Operating Activity',
                      'Cash From Operating activities']
        capex_keys = ['Capital Expenditures',
                      'Capital Expenditure',
                      'Purchase of property, plant and equipment',
                      'Capital Expenditure Reported']

        op_cash_flow = None
        capex = None

        for key in op_cf_keys:
            if key in cf_df.index:
                op_cash_flow = cf_df.loc[key].dropna()
                break
        for key in capex_keys:
            if key in cf_df.index:
                capex = cf_df.loc[key].dropna()
                break

        if op_cash_flow is None or capex is None:
            raise ValueError("Could not find Operating Cash Flow or Capital Expenditures rows.")

        common_dates = op_cash_flow.index.intersection(capex.index)
        common_dates = common_dates.sort_values(ascending=False)[:years]

        historical_fcf = op_cash_flow[common_dates] - capex[common_dates]
        historical_fcf = historical_fcf[::-1]

        last_fcf = historical_fcf.iloc[-1]

        projected_fcfs = [last_fcf * ((1 + growth_rate) ** t) for t in range(1, years + 1)]
        pv_fcfs = [fcf / ((1 + discount_rate) ** t) for t, fcf in enumerate(projected_fcfs, 1)]

        terminal_value = (projected_fcfs[-1] * (1 + terminal_growth)) / (discount_rate - terminal_growth)
        pv_terminal = terminal_value / ((1 + discount_rate) ** years)

        enterprise_value = sum(pv_fcfs) + pv_terminal

        years_list = [f"Year {i}" for i in range(1, years + 1)]
        dcf_summary = pd.DataFrame({
            "Projected FCF (₹)": projected_fcfs,
            "Discount Factor": [(1 + discount_rate) ** t for t in range(1, years + 1)],
            "Present Value of FCF (₹)": pv_fcfs,
        }, index=years_list)

        terminal_row = pd.DataFrame({
            "Projected FCF (₹)": [""],
            "Discount Factor": [""],
            "Present Value of FCF (₹)": [pv_terminal]
        }, index=["Terminal Value"])

        dcf_summary = pd.concat([dcf_summary, terminal_row])

        sum_row = pd.DataFrame({
            "Projected FCF (₹)": [""],
            "Discount Factor": [""],
            "Present Value of FCF (₹)": [enterprise_value]
        }, index=["Enterprise Value"])

        dcf_summary = pd.concat([dcf_summary, sum_row])

        return dcf_summary, enterprise_value

    except Exception as e:
        raise RuntimeError(f"DCF calculation failed: {e}")

# === Plot Price & Indicators ===
def plot_graph_save(df, ticker, scale_factor=1, price_unit='₹'):
    plt.figure(figsize=(14, 8))
    plt.subplot(2, 1, 1)
    plt.plot(df['Close'] / scale_factor, label=f'Close Price ({price_unit})')
    plt.plot(df['MA50'] / scale_factor, label=f'50-day MA ({price_unit})')
    plt.plot(df['MA200'] / scale_factor, label=f'200-day MA ({price_unit})')
    plt.title(f'{ticker} Price and Moving Averages')
    plt.xlabel('Date')
    plt.ylabel(f'Price ({price_unit})')
    plt.legend()
    plt.grid(True)

    plt.subplot(2, 1, 2)
    plt.plot(df['RSI'], label='RSI', color='orange')
    plt.axhline(70, color='red', linestyle='--', label='Overbought (70)')
    plt.axhline(30, color='green', linestyle='--', label='Oversold (30)')
    plt.title(f'{ticker} RSI')
    plt.xlabel('Date')
    plt.ylabel('RSI')
    plt.legend()
    plt.grid(True)

    plt.tight_layout()
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    plt.savefig(tmp_file.name, format='png')
    plt.close()
    return tmp_file.name

# === Excel Formatting Helpers ===
def format_sheet_headers(ws):
    fill = PatternFill("solid", fgColor="4F81BD")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_currency_format(ws, cols):
    fmt = '"₹"#,##0.00'
    for c in cols:
        col_letter = ws.cell(row=1, column=c).column_letter
        for cell in ws[col_letter][1:]:
            if cell.value is not None:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal='right')

def zebra_striping(ws, start=2):
    fill1 = PatternFill("solid", fgColor="F4F6F6")
    fill2 = PatternFill("solid", fgColor="FFFFFF")
    for i, row in enumerate(ws.iter_rows(min_row=start), start=start):
        fill = fill1 if i % 2 == 0 else fill2
        for cell in row:
            cell.fill = fill

def add_thin_border(ws):
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

def autofit_columns(ws):
    for col in ws.columns:
        max_val = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_val:
                    max_val = length
        ws.column_dimensions[col_letter].width = max_val + 3

def conditional_format_rsi(ws, col_letter):
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ws.conditional_formatting.add(f'{col_letter}2:{col_letter}1048576', CellIsRule(operator='greaterThan', formula=['70'], fill=red_fill))
    ws.conditional_formatting.add(f'{col_letter}2:{col_letter}1048576', CellIsRule(operator='lessThan', formula=['30'], fill=green_fill))

def color_investment_decision(ws):
    cell = ws['G2']
    val = cell.value
    if val == 'BUY':
        cell.font = Font(bold=True, color='006100')
        cell.fill = PatternFill('solid', fgColor='C6EFCE')
    elif val == 'SELL':
        cell.font = Font(bold=True, color='9C0006')
        cell.fill = PatternFill('solid', fgColor='FFC7CE')
    elif val == 'HOLD':
        cell.font = Font(bold=True, color='1F497D')
        cell.fill = PatternFill('solid', fgColor='D9E1F2')

def highlight_important_rows(ws, terms):
    bold = Font(bold=True, color='000000')
    fill = PatternFill('solid', fgColor='FFF2CC')
    for row in ws.iter_rows(min_row=2):
        first_cell = row[0]
        if first_cell.value and any(term.lower() in str(first_cell.value).lower() for term in terms):
            for cell in row:
                cell.font = bold
                cell.fill = fill

def format_analysis(ws, unit_label):
    format_sheet_headers(ws)
    apply_currency_format(ws, [2, 3, 4])
    zebra_striping(ws)
    add_thin_border(ws)
    autofit_columns(ws)
    conditional_format_rsi(ws, 'E')
    color_investment_decision(ws)
    ws.freeze_panes = 'A2'
    # Add unit annotation row
    ws.insert_rows(1)
    ws['A1'] = f"* All price values are in {unit_label}"
    ws['A1'].font = Font(italic=True, color='666666')

def format_financial(ws, important_terms, unit_label):
    format_sheet_headers(ws)
    money_cols = list(range(2, ws.max_column + 1))
    apply_currency_format(ws, money_cols)
    zebra_striping(ws)
    add_thin_border(ws)
    autofit_columns(ws)
    ws.freeze_panes = 'A2'
    highlight_important_rows(ws, important_terms)
    ws.insert_rows(1)
    ws['A1'] = f"* All values are in {unit_label}"
    ws['A1'].font = Font(italic=True, color='666666')

def format_dcf(ws, unit_label):
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = PatternFill('solid', fgColor='C6D9F1')
    ws['A2'].font = Font(italic=True, color='666666')

    header_row = 4
    for cell in ws[header_row]:
        cell.fill = PatternFill('solid', fgColor='4F81BD')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    apply_currency_format(ws, [2, 3, 4])
    zebra_striping(ws, start=header_row + 1)
    add_thin_border(ws)
    autofit_columns(ws)
    highlight_important_rows(ws, ['Terminal Value', 'Enterprise Value'])
    ws.freeze_panes = f'A{header_row + 1}'
    # Add unit annotation above all
    ws.insert_rows(1)
    ws['A1'] = f"* All values are in {unit_label}"
    ws['A1'].font = Font(italic=True, color='666666')

# === Export to Excel ===
def export_to_excel(df_analysis, ticker, decision, plot_path, bs, inc, cf, dcf_df, dcf_val, market_cap, val_msg, filename):
    wb = Workbook()

    # Determine scaling & unit label for analysis prices
    scale_analysis, unit_analysis = determine_scale_and_label(df_analysis[['Close', 'MA50', 'MA200']])
    df_analysis_scaled = df_analysis.copy()
    for col in ['Close', 'MA50', 'MA200']:
        df_analysis_scaled[col] = df_analysis[col] / scale_analysis

    ws_analysis = wb.active
    ws_analysis.title = "Analysis"
    for row in dataframe_to_rows(df_analysis_scaled[['Close', 'MA50', 'MA200', 'RSI']], index=True, header=True):
        ws_analysis.append(row)
    ws_analysis['G1'] = "Investment Decision:"
    ws_analysis['G2'] = decision
    img = XLImage(plot_path)
    img.anchor = 'I5'
    ws_analysis.add_image(img)
    format_analysis(ws_analysis, unit_analysis)

    def scale_financials_to_crores(df):
        if df is None:
            return None, 1, "₹"
        scale, unit = determine_scale_and_label(df)
        df_scaled = df.copy()
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                df_scaled[col] = df[col] / scale
        return df_scaled, scale, unit

    bs_scaled, _, bs_unit = scale_financials_to_crores(bs)
    inc_scaled, _, inc_unit = scale_financials_to_crores(inc)
    cf_scaled, _, cf_unit = scale_financials_to_crores(cf)

    bs_terms = ["Total Assets", "Total Liab", "Total Equity", "Cash", "Inventory", "Long Term Debt"]
    inc_terms = ["Total Revenue", "Gross Profit", "Operating Income", "Net Income"]
    cf_terms = ["Operating Cash Flow", "Capital Expenditures", "Free Cash Flow", "Net Change in Cash"]

    if bs_scaled is not None and not bs_scaled.empty:
        ws_bs = wb.create_sheet("Balance Sheet")
        ws_bs.append([f"* All values are in {bs_unit}"])
        ws_bs.append([])
        for row in dataframe_to_rows(bs_scaled, index=True, header=True):
            ws_bs.append(row)
        format_financial(ws_bs, bs_terms, bs_unit)
    else:
        print(f"No Balance Sheet data for {ticker}.")

    if inc_scaled is not None and not inc_scaled.empty:
        ws_inc = wb.create_sheet("Income Statement")
        ws_inc.append([f"* All values are in {inc_unit}"])
        ws_inc.append([])
        for row in dataframe_to_rows(inc_scaled, index=True, header=True):
            ws_inc.append(row)
        format_financial(ws_inc, inc_terms, inc_unit)
    else:
        print(f"No Income Statement data for {ticker}.")

    if cf_scaled is not None and not cf_scaled.empty:
        ws_cf = wb.create_sheet("Cash Flow")
        ws_cf.append([f"* All values are in {cf_unit}"])
        ws_cf.append([])
        for row in dataframe_to_rows(cf_scaled, index=True, header=True):
            ws_cf.append(row)
        format_financial(ws_cf, cf_terms, cf_unit)
    else:
        print(f"No Cash Flow data for {ticker}.")

    # Scale & label for DCF sheet data
    if not dcf_df.empty:
        dcf_numeric_cols = dcf_df.select_dtypes(include=[float, int]).columns
        max_dcf_val = dcf_df[dcf_numeric_cols].abs().max().max()
    else:
        max_dcf_val = 0

    if max_dcf_val >= 1e7:
        dcf_scale = 1e7
        dcf_unit = "Crores INR (₹)"
    elif max_dcf_val >= 1e5:
        dcf_scale = 1e5
        dcf_unit = "Lakhs INR (₹)"
    else:
        dcf_scale = 1
        dcf_unit = "INR (₹)"

    dcf_scaled = dcf_df.copy()
    for col in dcf_numeric_cols:
        dcf_scaled[col] = dcf_scaled[col] / dcf_scale

    dcf_val_scaled = dcf_val / dcf_scale if isinstance(dcf_val, (int, float)) else dcf_val
    mc_display = market_cap / dcf_scale if isinstance(market_cap, (int, float)) else market_cap

    ws_dcf = wb.create_sheet("DCF Analysis")
    ws_dcf.append(["Discounted Cash Flow Valuation Summary"])
    ws_dcf.append([f"* All values are in {dcf_unit}"])
    ws_dcf.append([])
    for row in dataframe_to_rows(dcf_scaled, index=True, header=True):
        ws_dcf.append(row)
    ws_dcf.append([])
    ws_dcf.append(["Enterprise Value (₹, scaled)", dcf_val_scaled])
    ws_dcf.append(["Market Capitalization (₹, scaled)", mc_display])
    ws_dcf.append([])
    ws_dcf.append([val_msg])
    last_row = ws_dcf.max_row
    cell = ws_dcf[f'A{last_row}']

    val_msg_lower = val_msg.lower()
    if "undervalued" in val_msg_lower:
        cell.font = Font(bold=True, color="006100")
    elif "overvalued" in val_msg_lower:
        cell.font = Font(bold=True, color="9C0006")
    else:
        cell.font = Font(bold=True, color="000000")

    format_dcf(ws_dcf, dcf_unit)

    wb.save(filename)
    print(f"Excel saved: {os.path.abspath(filename)}")

    if os.path.exists(plot_path):
        os.unlink(plot_path)

# === Main ===
def main():
    ticker = input("Enter stock symbol without suffix (e.g., RELIANCE, TCS): ").strip().upper()
    exchange = input("Is the stock listed on NSE or BSE? Enter NSE or BSE: ").strip().upper()

    if exchange == "NSE":
        ticker_full = ticker + ".NS"
    elif exchange == "BSE":
        ticker_full = ticker + ".BO"
    else:
        print("Invalid exchange input; choose NSE or BSE.")
        return

    try:
        data = fetch_data(ticker_full)
        data = calculate_indicators(data)
        data = data.dropna(subset=['MA50', 'MA200', 'RSI'])
        if data.empty:
            raise Exception("Insufficient data points for indicator calculation.")

        decision = investment_decision(data)
        print(f"Investment Decision for {ticker_full}: {decision}")
        print("\nLatest indicator values:")
        print(data[['Close','MA50','MA200','RSI']].tail())

        plot_path = plot_graph_save(data, ticker_full, scale_factor=1, price_unit='₹')

        bs, inc, cf = fetch_financial_statements(ticker_full)

        try:
            dcf_df, dcf_val = perform_dcf_analysis(cf)
            print(f"DCF Enterprise Value (₹): {dcf_val:,.2f}")
        except Exception as e:
            print(f"DCF calculation failed: {e}")
            dcf_df = pd.DataFrame()
            dcf_val = "N/A"

        market_cap_raw = get_market_cap(ticker_full)
        if market_cap_raw:
            market_cap_crores = market_cap_raw
        else:
            market_cap_crores = "N/A"

        if dcf_val != "N/A" and market_cap_crores != "N/A" and isinstance(market_cap_crores, (int,float)):
            if dcf_val > market_cap_crores:
                valuation_message = "According to DCF valuation, the company appears undervalued."
            elif dcf_val < market_cap_crores:
                valuation_message = "According to DCF valuation, the company appears overvalued."
            else:
                valuation_message = "According to DCF valuation, the company appears fairly valued."
        else:
            valuation_message = "Insufficient data to determine valuation status."

        excel_file = f"{ticker}_analysis.xlsx"
        export_to_excel(data, ticker_full, decision, plot_path, bs, inc, cf, dcf_df, dcf_val, market_cap_crores, valuation_message, excel_file)

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
